import os
import openpyxl
from slackbot.bot import Bot
from slackbot.bot import respond_to
from slackbot.bot import listen_to
from datetime import datetime

def make_file():
	wb = openpyxl.Workbook()
	ws = wb.worksheets[0]
	rng = ws["A1:D1"][0]
	rng[0].value = "DATE"
	rng[1].value = "USER_ID"
	rng[2].value = "TIME"
	wb.save("work_log.xlsx")

def get_day_of_week_jp(dt):
	w_list = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
	return(w_list[dt.weekday()])

@respond_to('使い方')
def how_rep(msg):
	msg.send('チャンネル内、もしくはリプライで「出勤」と入力することで出勤日と出勤時間を記録します。帰宅する際には「退勤」と入力することで退勤時間を記録します。\nもし、帰宅が24時以降になった場合は「日付」と「時間」が「〜」で結合されて記録されます。\n必ず「出勤」を入力してから「退勤」を入力してください。「出勤」を忘れた場合は管理者（吉田泰彦）に一報ください。')

@respond_to('出勤')
def punch_in(msg):
	log = datetime.now()
	wb = openpyxl.load_workbook('work_log.xlsx')
	ws = wb.worksheets[0]
	row_max = ws.max_row
	string = '{} [{}]'.format(log.strftime('%Y/%m/%d') , get_day_of_week_jp(log))
	col_date = [i for i, value in enumerate(ws['A']) if ws['A'][i].value == string]
	col_user = [i for i, value in enumerate(ws['B']) if ws['B'][i].value == '{}'.format(msg.body['user'])]

	if col_user:
		if ws[col_user[-1] + 1][0].value != string:
			ws[row_max + 1][0].value = '{} [{}]'.format(log.strftime('%Y/%m/%d') , get_day_of_week_jp(log))
			ws[row_max + 1][1].value = '{}'.format(msg.body['user'])
			ws[row_max + 1][2].value = '{}'.format(log.strftime('%H:%M')) + ' ～ '
			print('{} [{}]'.format(log.strftime('%Y/%m/%d') , get_day_of_week_jp(log)))
			print('{}'.format(msg.body['user']))
			print('{}'.format(log.strftime('%H:%M')) + ' ～ ')
	else: 
		ws[row_max + 1][0].value = '{} [{}]'.format(log.strftime('%Y/%m/%d') , get_day_of_week_jp(log))
		ws[row_max + 1][1].value = '{}'.format(msg.body['user'])
		ws[row_max + 1][2].value = '{}'.format(log.strftime('%H:%M')) + ' ～ '
		print('{} [{}]'.format(log.strftime('%Y/%m/%d') , get_day_of_week_jp(log)))
		print('{}'.format(msg.body['user']))
		print('{}'.format(log.strftime('%H:%M')) + ' ～ ')
	msg.send('出勤登録完了です。')
	wb.save('work_log.xlsx')

@respond_to('退勤')
def punch_in(msg):
	log = datetime.now()
	wb = openpyxl.load_workbook('work_log.xlsx')
	ws = wb.worksheets[0]
	row_max = ws.max_row

	string = '{} [{}]'.format(log.strftime('%Y/%m/%d') , get_day_of_week_jp(log))
	col_date = [i for i, value in enumerate(ws['A']) if ws['A'][i].value == string]
	col_user = [i for i, value in enumerate(ws['B']) if ws['B'][i].value == '{}'.format(msg.body['user'])]

	if col_user:
		if ws[col_user[-1] + 1][0].value == string:
			tmp = '{}'.format(ws[col_user[-1] + 1][2].value)
			ws[col_user[-1] + 1][2].value = tmp.split(' ～ ')[0] + ' ～ ' + log.strftime('%H:%M')
			msg.send('退勤登録完了です。')
			print('{} [{}]'.format(log.strftime('%Y/%m/%d') , get_day_of_week_jp(log)))
			print('{}'.format(msg.body['user']))
			print('{}'.format(log.strftime('%H:%M')))
		elif ws[col_user[-1] + 1][0].value != string:
			ws[col_user[-1] + 1][0].value = ws[col_user[-1] + 1][0].value + ' ～ ' + '{} [{}]'.format(log.strftime('%Y/%m/%d') , get_day_of_week_jp(log))
			tmp = '{}'.format(ws[col_user[-1] + 1][2].value)
			ws[col_user[-1] + 1][2].value = tmp.split(' ～ ')[0] + ' ～ ' + log.strftime('%H:%M')
			msg.send('退勤登録完了です。(当日帰宅の場合は「出勤」か「退勤」の登録にミスがあるので Excel ファイルの修正をお願いします。)')
			print('{} [{}]'.format(log.strftime('%Y/%m/%d') , get_day_of_week_jp(log)))
			print('{}'.format(msg.body['user']))
			print('{}'.format(log.strftime('%H:%M')))
	wb.save('work_log.xlsx')

def main():
	if (not os.path.isfile("work_log.xlsx")): make_file()
	bot = Bot()
	bot.run()

if __name__=='__main__':
	main()
